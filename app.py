import streamlit as st
import pandas as pd
import openpyxl
from io import BytesIO

# Импорты для генерации PDF
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

st.set_page_config(page_title="Финансовый ИИ-Агент", layout="wide")

st.title("🚗 Финансовый Автономный Агент Дилерского Центра")
st.write("Анализ влияния первичных статей расходов (исключая цветные суммирующие строки отделов) на общий бюджет.")

# Панель настроек в боковой панели
with st.sidebar:
    st.header("⚙️ Настройки анализа")
    target_column = st.text_input("Название столбца со статьями расходов:", value="Статья расходов")
    value_column = st.text_input("Название столбца со значениями (суммами):", value="Всего расходы")
    header_row = st.number_input("Строка с заголовками (в Excel нумерация с 1):", min_value=1, value=2)
    total_row_name = st.text_input("Название строки общего итога:", value="Всего по ДЦ")
    st.caption("ℹ️ Алгоритм автоматически исключает из ТОП-10 строки, имеющие цветовую заливку.")

# Блок загрузки файлов
col1, col2 = st.columns(2)
with col1:
    old_file = st.file_uploader("📂 Загрузите СТАРЫЙ отчет (прошлый месяц)", type=["xlsx"])
with col2:
    new_file = st.file_uploader("📂 Загрузите НОВЫЙ отчет (текущий месяц)", type=["xlsx"])

def is_colored(cell):
    """Проверяет, есть ли у ячейки цветная заливка (игнорирует белый/прозрачный)"""
    if cell.fill and cell.fill.fill_type:
        color = cell.fill.start_color.index
        if color and str(color) not in ['00000000', '0', 'FFFFFFFF', 'System_Color_Window']:
            return True
    return False

def generate_pdf(total_old, total_new, delta, df_top10):
    """Генерирует PDF-отчет в оперативной памяти (BytesIO)"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
    story = []
    
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'PDFTitle',
        parent=styles['Heading1'],
        fontSize=20,
        spaceAfter=15,
        textColor=colors.HexColor('#1E3A8A')
    )
    
    text_style = ParagraphStyle(
        'PDFText',
        parent=styles['Normal'],
        fontSize=11,
        spaceAfter=10
    )
    
    story.append(Paragraph("<b>Финансовый отчет Дилерского Центра</b>", title_style))
    story.append(Paragraph("Факторный анализ изменений в статьях расходов", text_style))
    story.append(Spacer(1, 15))
    
    story.append(Paragraph("<b>Общие финансовые результаты по ДЦ:</b>", styles['Heading2']))
    story.append(Paragraph(f"• Расходы за прошлый месяц: {total_old:,.2f} руб.", text_style))
    story.append(Paragraph(f"• Расходы за текущий месяц: {total_new:,.2f} руб.", text_style))
    story.append(Paragraph(f"• <b>Общее изменение расходов ДЦ: {delta:+,.2f} руб.</b>", text_style))
    story.append(Spacer(1, 15))
    
    story.append(Paragraph("<b>ТОП-10 главных изменений в статьях расходов:</b>", styles['Heading2']))
    story.append(Paragraph("В таблицу включены только чистые статьи расходов (цветные промежуточные итоги исключены).", text_style))
    story.append(Spacer(1, 5))
    
    table_data = [['№', 'Лист', 'Статья расходов', 'Было (руб.)', 'Стало (руб.)', 'Изменение', 'Доля в ДЦ']]
    
    for idx, row in df_top10.iterrows():
        table_data.append([
            str(idx),
            str(row['Лист']),
            str(row['Статья расходов']),
            f"{row['Было (руб.)']:,.2f}",
            f"{row['Стало (руб.)']:,.2f}",
            f"{row['Изменение (руб.)']:+,.2f}",
            f"{row['Доля во влиянии на общую разницу']:.2f}%"
        ])
        
    pdf_table = Table(table_data)
    
    pdf_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1E3A8A')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('ALIGN', (3,1), (-1,-1), 'RIGHT'),
        ('BOTTOMPADDING', (0,0), (-1,0), 8),
        ('BACKGROUND', (0,1), (-1,-1), colors.HexColor('#F3F4F6')),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E5E7EB')),
        ('FONTSIZE', (0,0), (-1,-1), 9),
    ]))
    
    story.append(pdf_table)
    doc.build(story)
    buffer.seek(0)
    return buffer

def parse_and_analyze():
    """Основная функция логики приложения, полностью изолированная для защиты от SyntaxError"""
    wb_old = openpyxl.load_workbook(old_file, data_only=True)
    wb_new = openpyxl.load_workbook(new_file, data_only=True)
    
    common_sheets = list(set(wb_old.sheetnames).intersection(set(wb_new.sheetnames)))
    
    if not common_sheets:
        st.error("❌ Ошибка: В файлах нет листов с одинаковыми названиями!")
        return
        
    all_expenses_changes = []
    total_old_dc = 0.0
    total_new_dc = 0.0
    total_row_found = False
    header_idx = int(header_row)
    
    for sheet_name in common_sheets:
        ws_old = wb_old[sheet_name]
        ws_new = wb_new[sheet_name]
        
        target_col_idx_old, value_col_idx_old = None, None
        target_col_idx_new, value_col_idx_new = None, None
        
        for col in range(1, ws_old.max_column + 1):
            val = str(ws_old.cell(row=header_idx, column=col).value).strip()
            if val == target_column: target_col_idx_old = col
            if val == value_column: value_col_idx_old = col
                
        for col in range(1, ws_new.max_column + 1):
            val = str(ws_new.cell(row=header_idx, column=col).value).strip()
            if val == target_column: target_col_idx_new = col
            if val == value_column: value_col_idx_new = col
        
        if target_col_idx_old and value_col_idx_old and target_col_idx_new and value_col_idx_new:
            dict_old = {}
            for r in range(header_idx + 1, ws_old.max_row + 1):
                cell_art = ws_old.cell(row=r, column=target_col_idx_old)
                cell_val = ws_old.cell(row=r, column=value_col_idx_old)
                
                if cell_art.value is not None:
                    art_str = str(cell_art.value).strip()
                    if art_str.lower() == total_row_name.lower().strip():
                        try: total_old_dc += float(cell_val.value or 0)
                        except: pass
                        total_row_found = True
                        continue
                    if is_colored(cell_art):
                        continue
                    dict_old[art_str] = cell_val.value

            dict_new = {}
            for r in range(header_idx + 1, ws_new.max_row + 1):
                cell_art = ws_new.cell(row=r, column=target_col_idx_new)
                cell_val = ws_new.cell(row=r, column=value_col_idx_new)
                
                if cell_art.value is not None:
                    art_str = str(cell_art.value).strip()
                    if art_str.lower() == total_row_name.lower().strip():
                        try: total_new_dc += float(cell_val.value or 0)
                        except: pass
                        continue
                    if is_colored(cell_art):
                        continue
                    dict_new[art_str] = cell_val.value
            
            sheet_articles = set(dict_old.keys()).union(set(dict_new.keys()))
            for article in sheet_articles:
                if article == "" or any(word in article.lower() for word in ["итого", "всего", "баланс", "результат", "свод"]):
                    continue
                try: val_old = float(dict_old.get(article, 0) or 0)
                except: val_old = 0.0
                try: val_new = float(dict_new.get(article, 0) or 0)
                except: val_new = 0.0
                
                item_delta = val_new - val_old
                abs_delta = abs(item_delta)
                
                if abs_delta > 0:
                    all_expenses_changes.append({
                        "Лист": sheet_name,
                        "Статья расходов": article,
                        "Было (руб.)": val_old,
                        "Стало (руб.)": val_new,
                        "Изменение (руб.)": item_delta,
                        "Абсолютное влияние (руб.)": abs_delta
                    })
    
    dc_delta = total_new_dc - total_old_dc
    
    st.subheader("📊 Общий финансовый результат по ДЦ")
    c1, c2, c3 = st.columns(3)
    with c1:
        st.metric("Расходы за прошлый месяц", f"{total_old_dc:,.2f} руб.")
    with c2:
        st.metric("Расходы за текущий месяц", f"{total_new_dc:,.2f} руб.")
    with c3:
        st.metric("Общее изменение расходов ДЦ", f"{dc_delta:+,.2f} руб.", delta_color="inverse")
        
    if not total_row_found:
        st.warning(f"⚠️ Строка '{total_row_name}' не найдена в файлах.")
    
    if all_expenses_changes:
        df_total_changes = pd.DataFrame(all_expenses_changes)
        top_10_changes = df_total_changes.sort_values(by="Абсолютное влияние (руб.)", ascending=False).head(10)
        
        if total_old_dc > 0:
            top_10_changes["Доля во влиянии на общую разницу"] = top_10_changes["Изменение (руб.)"] / total_old_dc * 100
        else:
            top_10_changes["Доля во влиянии на общую разницу"] = 0.0
        
        top_10_display = top_10_changes.drop(columns=["Абсолютное влияние (руб.)"], errors='ignore').reset_index(drop=True)
        top_10_display.index = top_10_display.index + 1
        
        st.subheader("📋 Директорский отчет: ТОП-10 чистых статей расходов")
        st.write("Суммирующие строки отделов отфильтрованы по цвету заливки. Показываются только прямые статьи расходов:")
        
        st.dataframe(
            top_10_display.style.format({
                "Было (руб.)": "{:,.2f}", 
                "Стало (руб.)": "{:,.2f}", 
                "Изменение (руб.)": "{:+,.2f}",
